import os
import openpyxl
from PyPDF2 import PdfMerger
import win32com.client
import tempfile
import shutil
import time
import gc

def process_excel_files_reestr():
    # Пути к файлам
    files_dir = "files"
    source_file = os.path.join(files_dir, "Reestr sobstvennikov.xlsx")
    template_file = os.path.join(files_dir, "Reestr dogovor.xlsx")
    output_dir = "output"
    
    # Создаем выходную директорию, если ее нет
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Создаем временную директорию для Excel файлов
    temp_dir = os.path.join(output_dir, "temp_reestr")
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    
    # Открываем исходный файл
    source_wb = openpyxl.load_workbook(source_file)
    source_ws = source_wb.active
    
    # Задаем фиксированный диапазон строк с 5 по 422 включительно
    start_row = 5
    end_row = 422
    row_count = end_row - start_row + 1
    
    print(f"[Reestr] Будет обработано {row_count} записей (строки с {start_row} по {end_row})")
    
    # Создаем список путей к созданным Excel файлам
    excel_files = []
    
    # Обрабатываем каждую строку
    for i in range(start_row, end_row + 1):
        # Получаем данные из исходного файла
        b_value = source_ws.cell(row=i, column=2).value  # B
        c_value = source_ws.cell(row=i, column=3).value  # C
        d_value = source_ws.cell(row=i, column=4).value  # D
        f_value = source_ws.cell(row=i, column=6).value  # F
        g_value = source_ws.cell(row=i, column=7).value  # G
        
        # Конвертируем этот файл напрямую в PDF с использованием win32com
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        
        try:
            # Открываем шаблон
            template_path = os.path.abspath(template_file)
            wb = excel.Workbooks.Open(template_path)
            
            # Заполняем лист
            ws = wb.Worksheets(1)  # индексация с 1 в COM
            
            # Устанавливаем значения в нужных ячейках
            try:
                ws.Range("G6").Value = b_value
                ws.Range("G14").Value = c_value
                ws.Range("G8").Value = d_value
                ws.Range("G10").Value = f_value
                ws.Range("G12").Value = g_value
            except Exception as e:
                print(f"Ошибка при заполнении данных: {e}")
            
            # Сохраняем промежуточный Excel файл
            output_excel = os.path.join(temp_dir, f"Reestr_dogovor_{i-4}.xlsx")
            wb.SaveAs(os.path.abspath(output_excel))
            excel_files.append(output_excel)
            
            # Экспортируем в PDF сразу
            pdf_path = os.path.join(output_dir, f"Reestr_dogovor_{i-4}.pdf")
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            
            wb.Close(False)
            print(f"[Reestr] Обработан файл для строки {i}")
            
        except Exception as e:
            print(f"Ошибка при обработке строки {i}: {e}")
        finally:
            excel.Quit()
            excel = None
            time.sleep(0.5)  # Небольшая задержка для освобождения ресурсов
            gc.collect()
    
    # Закрываем исходный файл
    source_wb.close()
    source_wb = None
    gc.collect()
    
    # Объединяем PDF-файлы
    pdf_files = [os.path.join(output_dir, f"Reestr_dogovor_{i-4}.pdf") for i in range(start_row, end_row + 1)]
    merged_pdf_path = os.path.join(output_dir, "Итоговый_реестр.pdf")
    
    try:
        merger = PdfMerger()
        
        for pdf in pdf_files:
            if os.path.exists(pdf):
                merger.append(pdf)
        
        merger.write(merged_pdf_path)
        merger.close()
        
        print(f"Создан итоговый PDF: {merged_pdf_path}")
        
        # Добавляем задержку перед удалением файлов
        time.sleep(1)
        
        # Удаляем отдельные PDF файлы
        for pdf in pdf_files:
            if os.path.exists(pdf):
                try:
                    os.remove(pdf)
                except Exception as e:
                    print(f"Ошибка при удалении файла {pdf}: {e}")
        
        # Удаляем временные Excel файлы
        for file in excel_files:
            if os.path.exists(file):
                try:
                    os.remove(file)
                except Exception as e:
                    print(f"Ошибка при удалении файла {file}: {e}")
        
        # Удаляем временную директорию
        if os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                print(f"Ошибка при удалении директории {temp_dir}: {e}")
    
    except Exception as e:
        print(f"Ошибка при объединении PDF: {e}")
    
    print("[Reestr] Обработка завершена успешно!")
    return merged_pdf_path

def process_excel_files_spisok():
    # Пути к файлам
    files_dir = "files"
    source_file = os.path.join(files_dir, "Reestr sobstvennikov.xlsx")
    template_file = os.path.join(files_dir, "Spisok.xlsx")
    output_dir = "output"
    
    # Создаем выходную директорию, если ее нет
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Создаем временную директорию для Excel файлов
    temp_dir = os.path.join(output_dir, "temp_spisok")
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    
    # Открываем исходный файл
    source_wb = openpyxl.load_workbook(source_file)
    source_ws = source_wb.active
    
    # Задаем фиксированный диапазон строк с 5 по 422 включительно
    start_row = 5
    end_row = 422
    row_count = end_row - start_row + 1
    
    print(f"[Spisok] Будет обработано {row_count} записей (строки с {start_row} по {end_row})")
    
    # Создаем список путей к созданным Excel файлам
    excel_files = []
    
    # Обрабатываем каждую строку
    for i in range(start_row, end_row + 1):
        # Получаем данные из исходного файла
        b_value = source_ws.cell(row=i, column=2).value  # B
        c_value = source_ws.cell(row=i, column=3).value  # C
        f_value = source_ws.cell(row=i, column=6).value  # F
        g_value = source_ws.cell(row=i, column=7).value  # G
        
        # Конвертируем этот файл напрямую в PDF с использованием win32com
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        
        try:
            # Открываем шаблон
            template_path = os.path.abspath(template_file)
            wb = excel.Workbooks.Open(template_path)
            
            # Заполняем лист
            ws = wb.Worksheets(1)  # индексация с 1 в COM
            
            # Устанавливаем значения в нужных ячейках
            try:
                ws.Range("C10").Value = b_value
                ws.Range("B10").Value = c_value
                ws.Range("E10").Value = f_value
                ws.Range("D10").Value = g_value
            except Exception as e:
                print(f"Ошибка при заполнении данных: {e}")
            
            # Сохраняем промежуточный Excel файл
            output_excel = os.path.join(temp_dir, f"Spisok_{i-4}.xlsx")
            wb.SaveAs(os.path.abspath(output_excel))
            excel_files.append(output_excel)
            
            # Экспортируем в PDF сразу
            pdf_path = os.path.join(output_dir, f"Spisok_{i-4}.pdf")
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            
            wb.Close(False)
            print(f"[Spisok] Обработан файл для строки {i}")
            
        except Exception as e:
            print(f"Ошибка при обработке строки {i}: {e}")
        finally:
            excel.Quit()
            excel = None
            time.sleep(0.5)  # Небольшая задержка для освобождения ресурсов
            gc.collect()
    
    # Закрываем исходный файл
    source_wb.close()
    source_wb = None
    gc.collect()
    
    # Объединяем PDF-файлы
    pdf_files = [os.path.join(output_dir, f"Spisok_{i-4}.pdf") for i in range(start_row, end_row + 1)]
    merged_pdf_path = os.path.join(output_dir, "Итоговый_список.pdf")
    
    try:
        merger = PdfMerger()
        
        for pdf in pdf_files:
            if os.path.exists(pdf):
                merger.append(pdf)
        
        merger.write(merged_pdf_path)
        merger.close()
        
        print(f"Создан итоговый PDF: {merged_pdf_path}")
        
        # Добавляем задержку перед удалением файлов
        time.sleep(1)
        
        # Удаляем отдельные PDF файлы
        for pdf in pdf_files:
            if os.path.exists(pdf):
                try:
                    os.remove(pdf)
                except Exception as e:
                    print(f"Ошибка при удалении файла {pdf}: {e}")
        
        # Удаляем временные Excel файлы
        for file in excel_files:
            if os.path.exists(file):
                try:
                    os.remove(file)
                except Exception as e:
                    print(f"Ошибка при удалении файла {file}: {e}")
        
        # Удаляем временную директорию
        if os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                print(f"Ошибка при удалении директории {temp_dir}: {e}")
    
    except Exception as e:
        print(f"Ошибка при объединении PDF: {e}")
    
    print("[Spisok] Обработка завершена успешно!")
    return merged_pdf_path

def process_excel_files_reshenie():
    # Пути к файлам
    files_dir = "files"
    source_file = os.path.join(files_dir, "Reestr sobstvennikov.xlsx")
    template_file = os.path.join(files_dir, "Reshenie.xlsx")
    output_dir = "output"
    
    # Создаем выходную директорию, если ее нет
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Создаем временную директорию для Excel файлов
    temp_dir = os.path.join(output_dir, "temp_reshenie")
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    
    # Открываем исходный файл
    source_wb = openpyxl.load_workbook(source_file)
    source_ws = source_wb.active
    
    # Задаем фиксированный диапазон строк с 5 по 422 включительно
    start_row = 5
    end_row = 422
    row_count = end_row - start_row + 1
    
    print(f"[Reshenie] Будет обработано {row_count} записей (строки с {start_row} по {end_row})")
    
    # Создаем список путей к созданным Excel файлам
    excel_files = []
    
    # Обрабатываем каждую строку
    for i in range(start_row, end_row + 1):
        # Получаем данные из исходного файла
        b_value = source_ws.cell(row=i, column=2).value  # B
        c_value = source_ws.cell(row=i, column=3).value  # C
        d_value = source_ws.cell(row=i, column=4).value  # D
        e_value = source_ws.cell(row=i, column=5).value  # E
        f_value = source_ws.cell(row=i, column=6).value  # F
        g_value = source_ws.cell(row=i, column=7).value  # G
        
        # Конвертируем этот файл напрямую в PDF с использованием win32com
        # Это обойдет проблему с объединенными ячейками
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        
        try:
            # Открываем шаблон
            template_path = os.path.abspath(template_file)
            wb = excel.Workbooks.Open(template_path)
            
            # Заполняем первый лист
            ws1 = wb.Worksheets(1)  # индексация с 1 в COM
            
            # Устанавливаем значения в нужных ячейках
            # Используем другой формат для задания значений через COM
            try:
                ws1.Range("C6").Value = b_value
                ws1.Range("A4").Value = c_value
                ws1.Range("C9").Value = d_value
                ws1.Range("C10").Value = e_value
                ws1.Range("C11").Value = f_value
                ws1.Range("A7").Value = g_value
            except Exception as e:
                print(f"Ошибка при заполнении первого листа: {e}")
            
            # Заполняем второй лист
            ws2 = wb.Worksheets(2)  # индексация с 1 в COM
            
            try:
                ws2.Range("C6").Value = b_value
                ws2.Range("A4").Value = c_value
                ws2.Range("C9").Value = d_value
                ws2.Range("C10").Value = e_value
                ws2.Range("C11").Value = f_value
                ws2.Range("A7").Value = g_value
            except Exception as e:
                print(f"Ошибка при заполнении второго листа: {e}")
            
            # Сохраняем промежуточный Excel файл
            output_excel = os.path.join(temp_dir, f"Reshenie_{i-4}.xlsx")
            wb.SaveAs(os.path.abspath(output_excel))
            excel_files.append(output_excel)
            
            # Экспортируем в PDF сразу
            pdf_path = os.path.join(output_dir, f"Reshenie_{i-4}.pdf")
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            
            wb.Close(False)
            print(f"[Reshenie] Обработан файл для строки {i}")
            
        except Exception as e:
            print(f"Ошибка при обработке строки {i}: {e}")
        finally:
            excel.Quit()
            excel = None
            time.sleep(0.5)  # Небольшая задержка для освобождения ресурсов
            gc.collect()
    
    # Закрываем исходный файл
    source_wb.close()
    source_wb = None
    gc.collect()
    
    # Объединяем PDF-файлы
    pdf_files = [os.path.join(output_dir, f"Reshenie_{i-4}.pdf") for i in range(start_row, end_row + 1)]
    merged_pdf_path = os.path.join(output_dir, "Итоговое_решение.pdf")
    
    try:
        merger = PdfMerger()
        
        for pdf in pdf_files:
            if os.path.exists(pdf):
                merger.append(pdf)
        
        merger.write(merged_pdf_path)
        merger.close()
        
        print(f"Создан итоговый PDF: {merged_pdf_path}")
        
        # Добавляем задержку перед удалением файлов
        time.sleep(1)
        
        # Удаляем отдельные PDF файлы
        for pdf in pdf_files:
            if os.path.exists(pdf):
                try:
                    os.remove(pdf)
                except Exception as e:
                    print(f"Ошибка при удалении файла {pdf}: {e}")
        
        # Удаляем временные Excel файлы
        for file in excel_files:
            if os.path.exists(file):
                try:
                    os.remove(file)
                except Exception as e:
                    print(f"Ошибка при удалении файла {file}: {e}")
        
        # Удаляем временную директорию
        if os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                print(f"Ошибка при удалении директории {temp_dir}: {e}")
    
    except Exception as e:
        print(f"Ошибка при объединении PDF: {e}")
    
    print("[Reshenie] Обработка завершена успешно!")
    return merged_pdf_path

if __name__ == "__main__":
    print("=== Запуск обработки файлов ===")
    
    # Выберите, какие шаблоны нужно обработать:
    # Вариант 1: Только реестр договоров
    # reestr_pdf = process_excel_files_reestr()
    # spisok_pdf = None
    # reshenie_pdf = None
    
    # Вариант 2: Только список
    # reestr_pdf = None
    # spisok_pdf = process_excel_files_spisok()
    # reshenie_pdf = None
    
    # Вариант 3: Только решение
    reestr_pdf = None
    spisok_pdf = None
    reshenie_pdf = process_excel_files_reshenie()
    
    # Вариант 4: Все шаблоны (по умолчанию)
    #reestr_pdf = process_excel_files_reestr()
    #spisok_pdf = process_excel_files_spisok()
    #reshenie_pdf = process_excel_files_reshenie()
    
    print("=== Обработка завершена ===") 